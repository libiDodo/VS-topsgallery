import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re

# --- НАСТРОЙКИ ---
excel_file = 'Backend/Data/TOPS.xlsx'
json_file = 'Backend/Data/TOPS.json'
# Предполагаем, что столбец "К-во" существует
# -------------------------------------------------

# Чтение данных через pandas (для основного содержимого)
df = pd.read_excel(excel_file)

# Загружаем книгу через openpyxl для доступа к стилям
wb = load_workbook(excel_file, data_only=True)
ws = wb.active

# Определяем индекс столбца "К-во" (поиск по заголовку в первой строке)
col_index = None
for i, cell in enumerate(ws[1], start=1):
    if cell.value == "К-во":
        col_index = i
        break

if col_index is None:
    raise ValueError("Столбец 'К-во' не найден. Проверьте название столбца в Excel.")

# Собираем цвета для каждой строки (начиная со 2-й строки, т.к. 1-я — заголовок)
colors = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_index, max_col=col_index):
    cell = row[0]
    fill = cell.fill
    color_hex = None

    # Определяем тип заливки и получаем цвет
    if isinstance(fill, PatternFill) and fill.fill_type == 'solid':
        # Для сплошной заливки используем fgColor
        rgb = fill.fgColor.rgb if fill.fgColor else None
    else:
        # Альтернативно, можно использовать start_color (например, для градиентов, но редко)
        rgb = fill.start_color.rgb if hasattr(fill, 'start_color') else None

    if rgb is not None:
        # Если rgb — объект (например, <RGB 00FF00>), пытаемся получить строковое представление
        if hasattr(rgb, 'rgb'):
            rgb = rgb.rgb  # некоторые объекты имеют атрибут rgb
        rgb_str = str(rgb) if not isinstance(rgb, str) else rgb

        # Очищаем строку: убираем возможные префиксы и оставляем только HEX
        # Ожидаем строку вида 'FF00FF00' или '00FF00'
        rgb_str = rgb_str.strip().upper()
        # Проверяем, соответствует ли строка HEX-формату (6 или 8 символов)
        if re.match(r'^[A-F0-9]{6}$', rgb_str) or re.match(r'^[A-F0-9]{8}$', rgb_str):
            if len(rgb_str) == 8 and rgb_str.startswith('FF'):
                rgb_str = rgb_str[2:]  # убираем альфа-канал
            color_hex = rgb_str
        # Если не подходит, возможно это цвет из темы, пропускаем
    colors.append(color_hex)

# Добавляем столбец с цветом в DataFrame
df['color'] = colors

# Сохраняем в JSON с русскими символами
df.to_json(json_file, orient='records', force_ascii=False, indent=4, date_format='iso')

print(f"JSON сохранён в {json_file}")